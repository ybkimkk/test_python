import http.client
import json
import os
import random
import re
import time

import xlrd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

import util.deviceUtils as deviceUtils
from util import configUtils
from util.adbUtils import device_ip_path
from util.adbUtils import message_path
from util.adbUtils import user_path

config = configUtils.config


def save_device_ip(ip):
    # 文件路径
    file_path = 'data/device_ip.xlsx'

    # 检查文件是否存在
    if os.path.exists(file_path):
        # 如果文件存在，加载现有的工作簿
        wb = load_workbook(file_path)
        ws = wb.active  # 获取默认工作表
    else:
        wb = Workbook()
        ws = wb.active

    ip_found = False
    for row in ws.iter_rows(min_row=0):
        if row[0].value == ip:
            ip_found = True
            break

    if not ip_found:
        ws.append([ip])

    # 保存文件
    wb.save(file_path)


def get_device_ips():
    ip = xlrd.open_workbook(device_ip_path)
    ip_sheets = ip.sheets()[0]
    ip_list = []
    for row_idx in range(ip_sheets.nrows - 1, -1, -1):
        ip_list.append(ip_sheets.cell_value(row_idx, 0))
    return ip_list


def clear_ip():
    file_path = 'data/device_ip.xlsx'
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    wb.save(file_path)


def extract_ip(output):
    match = re.search(r'inet addr:(\d+\.\d+\.\d+\.\d+)', output)
    if match:
        return match.group(1)
    else:
        return None


def random_sleep():
    time.sleep(random.randint(2, int(config['sleepMax'])))


def get_user_list():
    user = xlrd.open_workbook(user_path)
    return user.sheets()[0]


def get_message():
    message = xlrd.open_workbook(message_path)
    message_sheets = message.sheets()[0]
    phrases_dict = {}
    # 读取 Excel 数据并按数字分类
    for row_idx in range(1, message_sheets.nrows):  # 从第2行开始读取（跳过表头）
        number = message_sheets.cell_value(row_idx, 0)  # 获取数字列（第一列）
        phrase = message_sheets.cell_value(row_idx, 1)  # 获取话术列（第二列）

        if number not in phrases_dict:
            phrases_dict[number] = []
        phrases_dict[number].append(phrase)

    # 随机选择一个数字
    random_num = random.choice(list(phrases_dict.keys()))

    # 根据选中的数字获取话术列表
    selected_phrases = phrases_dict[random_num]

    # 打印选中的数字和对应的话术
    return selected_phrases


def simulate_typo(text):
    text = list(text)  # 将字符串转为字符列表，便于修改
    length = len(text)

    for i in range(length):
        # 决定是否插入错误
        if random.random() < float(config['errMsgRate']):
            error_type = random.choice(["swap", "delete", "insert"])

            if error_type == "swap" and i < length - 1:  # 字符交换
                # 随机交换当前字符和下一个字符
                text[i], text[i + 1] = text[i + 1], text[i]
            elif error_type == "delete":  # 删除字符
                text[i] = ""  # 删除当前字符
            elif error_type == "insert":  # 插入字符
                random_char = random.choice("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789!@#$%^&*()")
                text.insert(i, random_char)
                i += 1
    return "".join(text)


def check_device():
    conn = http.client.HTTPConnection("localhost", 80)
    result = False
    try:
        # 请求头
        headers = {
            "Content-Type": "application/json",  # 指定内容类型为 JSON
        }
        # 请求体（参数）
        payload = {
            "deviceId": deviceUtils.get_machine_code()
        }
        conn.request("POST", "/api/checkDevice", body=json.dumps(payload), headers=headers)
        response = conn.getresponse()
        if response.status == 200:
            response_data = json.loads(response.read().decode())
            print(response_data.get('msg'))
            result = response_data.get('check')
            if not result:
                print("该设备已到期")
        else:
            print("服务器异常,请联系管理员")
            result = False
    except Exception as e:
        print(e)
    conn.close()
    return result
